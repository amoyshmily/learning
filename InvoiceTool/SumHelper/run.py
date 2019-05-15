"""
统计未开票汇总
2019-5-13 22:55:54

"""
from PyQt5.QtWidgets import QMainWindow, QTableWidget, QVBoxLayout, \
    QApplication, QAction, QTableWidgetItem, QAbstractItemView, QFileDialog
from PyQt5.QtGui import QIcon
import sys
import openpyxl as xl
import json


class SumHelper(QMainWindow):

    def __init__(self, parent=None):
        super(SumHelper, self).__init__(parent)
        self.table_widget = QTableWidget()
        self.tool_bar = None
        self.initUI()

    def initUI(self):
        self.setWindowTitle('张梦文的小助手 v1.0')
        self.setWindowIcon(QIcon('invoice.png'))
        self.resize(600, 400)

        # ------ 设置菜单栏 ------
        source_file = self.menuBar().addMenu('&文件')
        open_action = QAction(QIcon('excel.png'), '打开文件', self)
        open_action.setShortcut('Ctrl+O')
        open_action.triggered.connect(self.showOpenFileChooser)
        source_file.addAction(open_action)

        # ------ 设置工具栏 ------
        # 打开文件
        tool_open = self.addToolBar('Open')
        tool_open_action = QAction(QIcon('excel.png'), '选择数据源Excel文件', self)
        tool_open_action.triggered.connect(self.showOpenFileChooser)
        tool_open.addAction(tool_open_action)
        # 导出文件
        tool_save = self.addToolBar('Save')
        tool_save_action = QAction(QIcon('save.png'), '保存文件', self)
        tool_save_action.triggered.connect(self.showSaveFileChooser)
        tool_save.addAction(tool_save_action)

        # ------ 设置状态栏 ------
        self.statusBar().showMessage('程序已准备就绪')

        # 设置行数、列数
        self.table_widget.setRowCount(1)
        self.table_widget.setColumnCount(2)
        item_default0 = QTableWidgetItem("请选择Excel数据源文件")
        item_default1 = QTableWidgetItem("0")
        self.table_widget.setItem(0, 0, item_default0)
        self.table_widget.setItem(0, 1, item_default1)

        # 设置表头
        self.table_widget.setHorizontalHeaderLabels(['公司名称', '未开票合计'])

        # 禁止编辑
        self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # 整行选中
        # self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)

        # 内容自适应
        self.table_widget.resizeColumnsToContents()
        self.table_widget.resizeRowsToContents()

        # 控制表头的显示/隐藏
        self.table_widget.horizontalHeader().setVisible(True)
        self.table_widget.verticalHeader().setVisible(True)

        # 设置布局
        layout = QVBoxLayout()
        layout.addWidget(self.table_widget)
        self.setLayout(layout)
        self.setCentralWidget(self.table_widget)

    def reloadData(self, excel_file_name: str):
        """
        加载数据
        :param excel_file_name:
        :return:
        """

        print('正在加载数据....')
        data = self.getClientInfoAll(excel_file_name)
        print(data)

        # 重新设置行数
        row_num = self.getValidRowCount(data=data)
        self.table_widget.setRowCount(row_num)

        row_offset = 0
        for k in data.keys():

            client_name = data[k]['clientName']
            undo_total = data[k]['undoTotal']

            # 添加数据
            item0 = QTableWidgetItem(client_name)
            item1 = QTableWidgetItem(str(undo_total))

            # 如果未开票金额为0则不展示
            if undo_total:
                self.table_widget.setItem(row_offset, 0, item0)
                self.table_widget.setItem(row_offset, 1, item1)
            else:
                continue
            row_offset = row_offset+1

        # 内容自适应
        self.table_widget.resizeColumnsToContents()
        self.table_widget.resizeRowsToContents()

    @staticmethod
    def getValidRowCount(data: dict) -> int:
        """
        获取有效金额的行数
        :param data:数据
        :return:
        """
        count = 0
        for k in data.keys():
            if data[k]['undoTotal']:
                count = count+1
        return count

    def showOpenFileChooser(self):
        """
        打开文件对话框
        :return:
        """
        # TODO：过滤掉非excel的文件
        file_name = QFileDialog.getOpenFileName(self, '打开文件', 'd:/')

        if file_name[0]:
            self.statusBar().showMessage('数据加载中，请耐心等候......')
            self.reloadData(file_name[0])
            self.statusBar().showMessage('数据读取完毕！')

    def showSaveFileChooser(self):
        """
        保存文件对话框
        :return:
        """
        file_name = QFileDialog.getSaveFileName(self, '保存文件', 'd:/')
        if file_name[0]:
            # TODO:将记录写入csv文件
            with open(file_name[0], 'wb') as file:
                file.write('hello world'.encode(encoding='utf-8'))
            self.statusBar().showMessage('文件保存成功！')
        else:
            self.statusBar().showMessage('文件保存失败！')

    def calc(self, total):
        # 云和丰泽
        targets_dict = {'12SZHE1优': {'pcs': 735, 'price': 61},
                        '15HE优': {'pcs': 255, 'price': 75},
                        '18SZCE1优': {'pcs': 140, 'price': 95}}

        # 判断候选总金额应大于开票金额
        count = 0
        for v in targets_dict.values():
            count = count + v['pcs'] * v['price']

        if count < total:
            print('候选产品的总金额不足开票总金额，无法开票！')
            return

        print('指定目标开票金额：' + str(total) + '元')
        n = 0
        for i in range(1, 735):
            for j in range(1, 255):
                for k in range(1, 140):
                    if i * 61 + j * 75 + k * 95 == total:

                        print('成功配数方案：', '12SZHE1优=' + str(i) + '个,', '15HE优=' + str(j) + '个,',
                              '18SZCE1优=' + str(k) + '个。')

                        n = n + 1
                        if n == 2:
                            return

    def getClientInfoAll(self, excel_file: str) -> dict:
        """
        获取客户未开票数据
        :param excel_file: 数据源文件
        :return:客户信息
        """
        wbook = xl.load_workbook(excel_file, data_only=True)

        # 校验客户名称
        # if client_name in client_name_list:
        # 	print('目标客户实际存在。')
        # else:
        # 	print('查无此客户：“'+client_name+'”,请核实！')

        # 获取每个客户名下存在未开票的产品的信息（名称、未开票数量、单价）
        client_name_list = wbook.sheetnames

        clients_info = {}
        for i in range(len(client_name_list)):

            client_info = {}

            client_name = client_name_list[i]
            print('正在解析客户：', client_name, '的信息')
            sheet = wbook[client_name]

            # 获取产品名称列表
            product_dict = {}
            for cell in sheet['A']:
                if cell.value is None:
                    continue
                else:
                    if cell.row > 2:
                        product_dict[cell.value] = {'name': cell.value, 'row': cell.row}
            print('该客户名下共有' + str(len(product_dict.keys())) + '款产品。')

            # 获取总计所在列的列数索引
            cursor = 0
            for col in range(1, 100):
                cell = sheet.cell(row=2, column=col)
                if cell.value == '总计':
                    cursor = cell.column
            print('参考游标所在列数：' + str(cursor) + '列')

            if cursor == 0:
                print('工作表中未找到“总计”列，无法获取开票数量和产品单价信息，请核实！')
                continue

            # 获取未开票数量和总金额
            each_total = 0
            for val in product_dict.values():
                # 查询未开票数量
                undo_cell = sheet.cell(row=val['row'] + 1, column=cursor + 1)
                val['undoNum'] = undo_cell.value

                # 查询产品单价
                price_cell = sheet.cell(row=val['row'] + 1, column=cursor + 2)
                val['price'] = price_cell.value

                # 查询单个产品未开票合计
                val['undoSum'] = val['undoNum'] * val['price']

                # 以客户为单位，计算未开票总金额
                each_total = each_total + float(val['undoSum'])

            print(json.dumps(product_dict, ensure_ascii=False))

            client_info['clientName'] = client_name
            client_info['products'] = product_dict
            client_info['undoTotal'] = each_total

            clients_info['client' + str(i)] = client_info

        print(json.dumps(clients_info, ensure_ascii=False))
        return clients_info

    @staticmethod
    def getNotInvoicedSum(client_info_dict):
        sum_amount = 0
        for item in client_info_dict.values():
            sum_amount = sum_amount + item['undoSum']
        print(sum_amount)


def main():
    app = QApplication(sys.argv)
    win = SumHelper()
    win.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
