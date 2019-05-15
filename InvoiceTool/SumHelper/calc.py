import json
import openpyxl as xl


def calc(total):
    # 云和丰泽
    targets_dict = {'12SZHE1优': {'pcs': 735, 'price': 61},
                    '15HE优': {'pcs': 255, 'price': 75},
                    '18SZCE1优': {'pcs': 140, 'price': 95}}

    # 判断候选总金额应大于开票金额
    count = 0
    for v in targets_dict.values():
        count = count + v['pcs'] * v['price']

    if count < total:
        print('候选产品的总金额不足开票总金额，无法开票！')
        return

    print('指定目标开票金额：' + str(total) + '元')
    n = 0
    for i in range(1, 735):
        for j in range(1, 255):
            for k in range(1, 140):
                if i * 61 + j * 75 + k * 95 == total:

                    print('成功配数方案：', '12SZHE1优=' + str(i) + '个,', '15HE优=' + str(j) + '个,', '18SZCE1优=' + str(k) + '个。')

                    n = n + 1
                    if n == 2:
                        return


def getClientInfoAll(excel_file: str) -> dict:
    """
    获取客户未开票数据
    :param excel_file: 数据源文件
    :return:客户信息
    """
    wbook = xl.load_workbook(excel_file, data_only=True)

    # 校验客户名称
    # if client_name in client_name_list:
    # 	print('目标客户实际存在。')
    # else:
    # 	print('查无此客户：“'+client_name+'”,请核实！')

    # 获取每个客户名下存在未开票的产品的信息（名称、未开票数量、单价）
    client_name_list = wbook.sheetnames

    clients_info = {}
    for i in range(len(client_name_list)):

        client_info = {}

        client_name = client_name_list[i]
        print('正在解析客户：', client_name, '的信息')
        sheet = wbook[client_name]

        # 获取产品名称列表
        product_dict = {}
        for cell in sheet['A']:
            if cell.value is None:
                continue
            else:
                if cell.row > 2:
                    product_dict[cell.value] = {'name': cell.value, 'row': cell.row}
        print('该客户名下共有' + str(len(product_dict.keys())) + '款产品。')

        # 获取总计所在列的列数索引
        cursor = 0
        for col in range(1, 100):
            cell = sheet.cell(row=2, column=col)
            if cell.value == '总计':
                cursor = cell.column
        print('参考游标所在列数：' + str(cursor) + '列')

        if cursor == 0:
            print('工作表中未找到“总计”列，无法获取开票数量和产品单价信息，请核实！')
            continue

        # 获取未开票数量和总金额
        each_total = 0
        for val in product_dict.values():

            # 查询未开票数量
            undo_cell = sheet.cell(row=val['row'] + 1, column=cursor + 1)
            val['undoNum'] = undo_cell.value

            # 查询产品单价
            price_cell = sheet.cell(row=val['row'] + 1, column=cursor + 2)
            val['price'] = price_cell.value

            # 查询单个产品未开票合计
            val['undoSum'] = val['undoNum']*val['price']

            # 以客户为单位，计算未开票总金额
            each_total = each_total + float(val['undoSum'])

        print(json.dumps(product_dict, ensure_ascii=False))

        client_info['clientName'] = client_name
        client_info['products'] = product_dict
        client_info['undoTotal'] = each_total

        clients_info['client'+str(i)] = client_info

    print(json.dumps(clients_info, ensure_ascii=False))
    return clients_info


def getNotInvoicedSum(client_info_dict):
    sum_amount = 0
    for item in client_info_dict.values():
        sum_amount = sum_amount+item['undoSum']
    print(sum_amount)


if __name__ == '__main__':
    # calc(71970)

    file = 'invoice.xlsx'
    # client = '华发木业'

    d = getClientInfoAll(file)

    # getNotInvoicedSum(d)
